from selenium import webdriver
import time
import requests
import json
from openpyxl import Workbook
import hashlib
import os
import datetime
import sys

a=sys.argv[1]
b=sys.argv[2]
print(b)
print(a)

#启动webdriver
base_url ='https://www.toutiao.com'
brower =webdriver.Firefox()
brower.get(base_url)
brower.implicitly_wait(10)
brower.maximize_window() # 最大化窗口
brower.implicitly_wait(10)
brower.find_element_by_link_text(a).click()
brower.implicitly_wait(10)

title_list,url_list,sources_list=[],[],[]
contents_list=[]


def get_info():
    #利用xpath方法动态新闻标题信息
    titles= brower.find_elements_by_xpath('//div[@class="title-box"]/a')
    # 获取页面新闻标题并添加到列表中
    for title in titles:
        title_list.append(title.text)

    urls = brower.find_elements_by_xpath('//div[@class="title-box"]/a')
    for url in urls:
        url = url.get_attribute('href')
        url_list.append(url)
    sources = brower.find_elements_by_xpath('//a[@class="lbtn source"]')
    for source in sources:
        sources_list.append(source.text)

    contents = brower.find_elements_by_xpath('//div[@class="article-content"]')
    for content in contents:
        contents_list.append(content.text)

# 通过下拉进度条一直加载页面
def get_manyinfo():
    brower.execute_script("window.scrollTo(0,1000);")
    time.sleep(1)
    while len(title_list) < 30:
        for i in range(int(b)): #30
            brower.execute_script("window.scrollTo(0,document.body.scrollHeight);")
            time.sleep(3)
        get_info()
        brower.refresh()
    else:
        brower.close()


def savedata(title_list, url_list, sources_list,contents_list):  # 存储数据到文件
    # 存储数据到xlxs文件
    wb = Workbook()
    if not os.path.isdir(os.getcwd()+'/result'):   # 判断文件夹是否存在
        os.makedirs(os.getcwd()+'/result') # 新建存储文件夹
    filename = os.getcwd()+'/result/result-'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%m')+'.xlsx' # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'   # 更改工作表的标题
    ws['A1'] = '标题'   # 对表格加入标题
    ws['B1'] = '新闻链接'
    ws['C1'] = '头条号'
    ws['D1'] = '头条号链接'
    for row in range(2, len(title_list)+2):   # 将数据写入表格
        _= ws.cell(column=1, row=row, value=title_list[row-2])
        _= ws.cell(column=2, row=row, value=url_list[row-2])
        _= ws.cell(column=3, row=row, value=sources_list[row-2])
        _= ws.cell(column=4, row=row, value=contents_list[row-2])
    filename1='002.xlsx'
    wb.save(filename=filename)  # 保存文件
    wb.save(filename=filename1)  # 保存文件

def main():
    get_manyinfo()
    savedata(title_list, url_list, sources_list,contents_list)

if __name__=="__main__":
    main()
